"""
Excel handling for the RSFM website, in two parts:

1. LIVE MASTER WORKBOOKS — one .xlsx file per form (enquiries, applications)
   that lives on disk at DATA_DIR and gets a new row appended the instant
   someone submits the form. Open these files directly in Microsoft Excel
   (or any spreadsheet app) and every new submission shows up the next
   time the file is reloaded — no export step required.

2. ON-DEMAND FILTERED EXPORT — used by the admin panel's "Export to Excel"
   button, which builds a fresh workbook from whatever the admin currently
   has filtered/searched for in the dashboard.

Both write with the same professional formatting: bold navy header row,
zebra striping, frozen header, auto-filter, and sensible column widths.
"""
import io
import os
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("rsfm.excel")

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data", "exports")
ENQUIRIES_XLSX_PATH = os.path.join(DATA_DIR, "rsfm-enquiries-live.xlsx")
APPLICATIONS_XLSX_PATH = os.path.join(DATA_DIR, "rsfm-applications-live.xlsx")

ENQUIRY_HEADERS = [
    "ID", "Date & Time", "Full Name", "Phone Number", "Email Address",
    "Service Required", "Message", "Status",
]
APPLICATION_HEADERS = [
    "ID", "Date & Time", "Full Name", "Phone Number", "Email Address",
    "Position Applied For", "Experience", "Message", "Status",
]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="0E2447", end_color="0E2447", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Arial", size=10.5)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
STRIPE_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDE3EC"),
    right=Side(style="thin", color="DDE3EC"),
    top=Side(style="thin", color="DDE3EC"),
    bottom=Side(style="thin", color="DDE3EC"),
)


def _style_header_row(ws, headers):
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 26


def _style_data_row(ws, row_idx, num_cols):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGN
        cell.border = THIN_BORDER
    if row_idx % 2 == 0:
        for col_idx in range(1, num_cols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = STRIPE_FILL


def _apply_column_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _ensure_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def _create_workbook(path, sheet_title, headers, widths):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    _style_header_row(ws, headers)
    _apply_column_widths(ws, widths)
    ws.freeze_panes = "A2"
    wb.save(path)
    return wb


ENQUIRY_WIDTHS = {"A": 7, "B": 20, "C": 22, "D": 16, "E": 26, "F": 24, "G": 46, "H": 14}
APPLICATION_WIDTHS = {"A": 7, "B": 20, "C": 22, "D": 16, "E": 26, "F": 24, "G": 16, "H": 40, "I": 14}


def ensure_master_workbooks():
    """Create both live workbooks (with headers only) if they don't exist yet.
    Called once on app startup so the files are always present and openable."""
    _ensure_dir()
    if not os.path.exists(ENQUIRIES_XLSX_PATH):
        _create_workbook(ENQUIRIES_XLSX_PATH, "Enquiries", ENQUIRY_HEADERS, ENQUIRY_WIDTHS)
        logger.info("Created live enquiries workbook at %s", ENQUIRIES_XLSX_PATH)
    if not os.path.exists(APPLICATIONS_XLSX_PATH):
        _create_workbook(APPLICATIONS_XLSX_PATH, "Applications", APPLICATION_HEADERS, APPLICATION_WIDTHS)
        logger.info("Created live applications workbook at %s", APPLICATIONS_XLSX_PATH)


def _append_row(path, sheet_title, headers, widths, row_values):
    """Append one row to a live workbook, recreating it first if it's
    missing or corrupted. Never raises — a spreadsheet write should never
    prevent an enquiry/application from having already been saved to the
    database.
    """
    _ensure_dir()
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb[sheet_title] if sheet_title in wb.sheetnames else wb.active
        else:
            wb = _create_workbook(path, sheet_title, headers, widths)
            ws = wb.active

        next_row = ws.max_row + 1
        ws.append(row_values)
        _style_data_row(ws, next_row, len(headers))
        ws.auto_filter.ref = "A1:{0}{1}".format(get_column_letter(len(headers)), next_row)
        wb.save(path)
        return True
    except Exception as exc:  # noqa: BLE001 — log and continue, never break the request
        logger.error("Failed to append row to live workbook %s: %s", path, exc)
        return False


def append_enquiry_to_live_workbook(enquiry):
    row = [
        enquiry.id,
        enquiry.created_at.strftime("%d-%b-%Y %I:%M %p") if enquiry.created_at else "",
        enquiry.name or "",
        enquiry.phone or "",
        enquiry.email or "",
        enquiry.service or "",
        enquiry.message or "",
        enquiry.status or "",
    ]
    return _append_row(ENQUIRIES_XLSX_PATH, "Enquiries", ENQUIRY_HEADERS, ENQUIRY_WIDTHS, row)


def append_application_to_live_workbook(application):
    row = [
        application.id,
        application.created_at.strftime("%d-%b-%Y %I:%M %p") if application.created_at else "",
        application.name or "",
        application.phone or "",
        application.email or "",
        application.role or "",
        application.experience or "",
        application.message or "",
        application.status or "",
    ]
    return _append_row(
        APPLICATIONS_XLSX_PATH, "Applications", APPLICATION_HEADERS, APPLICATION_WIDTHS, row
    )


# ---------------------------------------------------------------------------
# On-demand filtered export (admin dashboard "Export to Excel" button)
# ---------------------------------------------------------------------------

def build_enquiries_workbook(enquiries):
    """Return an in-memory .xlsx (BytesIO) for a filtered/searched set of enquiries."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Enquiries"
    headers = ENQUIRY_HEADERS[1:]  # omit internal ID for the admin-facing export
    ws.append(headers)
    _style_header_row(ws, headers)

    for row_idx, enquiry in enumerate(enquiries, start=2):
        ws.append([
            enquiry.created_at.strftime("%d-%b-%Y %I:%M %p") if enquiry.created_at else "",
            enquiry.name or "",
            enquiry.phone or "",
            enquiry.email or "",
            enquiry.service or "",
            enquiry.message or "",
            enquiry.status or "",
        ])
        _style_data_row(ws, row_idx, len(headers))

    _apply_column_widths(ws, {"A": 20, "B": 22, "C": 16, "D": 26, "E": 24, "F": 46, "G": 14})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{0}{1}".format(get_column_letter(len(headers)), max(len(enquiries) + 1, 1))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_applications_workbook(applications):
    """Return an in-memory .xlsx (BytesIO) for a filtered/searched set of applications."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    headers = APPLICATION_HEADERS[1:]
    ws.append(headers)
    _style_header_row(ws, headers)

    for row_idx, application in enumerate(applications, start=2):
        ws.append([
            application.created_at.strftime("%d-%b-%Y %I:%M %p") if application.created_at else "",
            application.name or "",
            application.phone or "",
            application.email or "",
            application.role or "",
            application.experience or "",
            application.message or "",
            application.status or "",
        ])
        _style_data_row(ws, row_idx, len(headers))

    _apply_column_widths(ws, {"A": 20, "B": 22, "C": 16, "D": 26, "E": 24, "F": 16, "G": 40, "H": 14})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{0}{1}".format(get_column_letter(len(headers)), max(len(applications) + 1, 1))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
